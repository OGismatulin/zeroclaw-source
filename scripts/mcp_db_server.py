"""MCP DB Server — PostgreSQL access for ZeroClaw agent via MCP protocol."""

from __future__ import annotations

import json
import http.client
import os
import re
import shutil
import socket
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request

import psycopg2
from psycopg2 import errors as pg_errors

DDL_BLOCKLIST_RE = re.compile(
    r"\b(DROP|TRUNCATE|ALTER|CREATE|GRANT|REVOKE|COPY|VACUUM|REINDEX|CLUSTER)\b",
    re.IGNORECASE,
)

LIMIT_RE = re.compile(r"\bLIMIT\s+\d+", re.IGNORECASE)


def _strip_sql_noise(sql: str) -> str:
    """Remove SQL comments and string literals for safe keyword matching."""
    # Remove block comments
    sql = re.sub(r"/\*.*?\*/", "", sql, flags=re.DOTALL)
    # Remove line comments
    sql = re.sub(r"--[^\n]*", "", sql)
    # Remove single-quoted string literals (handles '' escapes)
    sql = re.sub(r"'(?:[^']|'')*'", "''", sql)
    # Remove dollar-quoted strings (PostgreSQL)
    sql = re.sub(r"\$\$.*?\$\$", "''", sql, flags=re.DOTALL)
    return sql


def check_sql_safety(sql: str) -> None:
    """Raise ValueError if SQL contains blocked DDL keywords."""
    cleaned = _strip_sql_noise(sql)
    match = DDL_BLOCKLIST_RE.search(cleaned)
    if match:
        raise ValueError(
            f"DDL statement blocked: '{match.group()}' is not allowed. "
            f"Only SELECT and DML (INSERT/UPDATE/DELETE) are permitted."
        )


def apply_auto_limit(sql: str) -> str:
    """Add LIMIT 1000 to SELECT/WITH queries that lack an explicit LIMIT."""
    stripped = sql.strip().rstrip(";")
    upper = stripped.upper()
    if not (upper.startswith("SELECT") or upper.startswith("WITH")):
        return sql
    if LIMIT_RE.search(stripped):
        return sql
    return stripped + " LIMIT 1000"


DATABASES: dict[str, dict[str, str]] = {
    # Catalog (prod: separate host slave.catalog.db)
    "catalog_microservice": {"description": "Ads, categories, micromarkets"},
    "global_catalog_microservice": {"description": "Global catalog data"},
    # Users & auth
    "user_microservice": {"description": "Users, profiles, auth"},
    "business_profile_microservice": {"description": "Business profiles"},
    # Payments & monetization
    "payment_microservice": {"description": "Payments, subscriptions"},
    "wallet_microservice": {"description": "Wallet, balance"},
    "tax_microservice": {"description": "Tax calculations"},
    # Sales & CRM
    "crm_microservice": {"description": "CRM leads, deals"},
    "tds_microservice": {"description": "TDS traffic distribution"},
    "campaign_microservice": {"description": "Campaign management"},
    "promo_microservice": {"description": "Promo codes, promotions"},
    # Content & moderation
    "moderation_microservice": {"description": "Content moderation"},
    "fraud_microservice": {"description": "Fraud detection"},
    "upload_microservice": {"description": "File uploads"},
    # Messaging & notifications
    "sender_microservice": {"description": "Message/notification sender"},
    "trigger_microservice": {"description": "Event triggers, automations"},
    # Search & discovery
    "search_microservice": {"description": "Search engine"},
    "marketplace_microservice": {"description": "Marketplace operations"},
    "micromarket_microservice": {"description": "Micromarkets, geo zones"},
    "location_microservice": {"description": "Locations, addresses"},
    # Analytics & experiments
    "analytics_microservice": {"description": "Analytics tracking"},
    "user_csm_analytics_microservice": {"description": "User CSM analytics"},
    "experiment_microservice": {"description": "A/B experiments"},
    # SEO & content
    "seo_microservice": {"description": "SEO optimization"},
    "page_microservice": {"description": "Static pages"},
    "translations_microservice": {"description": "Translations, i18n"},
    "social_feed_microservice": {"description": "Social feed"},
    "branding_microservice": {"description": "Brand profiles"},
    # Infrastructure
    "app_versioning_microservice": {"description": "App version management"},
}

# Databases that live on a separate prod host (slave.catalog.db)
PROD_CATALOG_HOST_DBS = {
    "catalog_microservice",
    "global_catalog_microservice",
    "bot_microservice",
    "catalog_microservice_history",
    "analytics",
}


def resolve_connection_params(env: str, db: str) -> dict:
    """Build psycopg2 connection params for the given environment and database."""
    if env not in ("stage", "prod"):
        raise ValueError(f"Invalid environment: '{env}'. Must be 'stage' or 'prod'.")
    if db not in DATABASES:
        raise ValueError(
            f"Unknown database: '{db}'. Available: {', '.join(sorted(DATABASES.keys()))}"
        )

    if env == "stage":
        return {
            "host": os.environ.get("PG_STAGE_HOST", "stage.postgres.kube-two.yallasvc.net"),
            "port": int(os.environ.get("PG_STAGE_PORT", "6432")),
            "user": os.environ.get("PG_STAGE_USER", "postgres"),
            "password": os.environ.get("PG_STAGE_PASSWORD", ""),
            "dbname": db,
            "sslmode": "disable",
        }

    # prod
    if db in PROD_CATALOG_HOST_DBS:
        host = os.environ.get("PG_PROD_CATALOG_HOST", "slave.catalog.db.yallasvc.net")
    else:
        host = os.environ.get("PG_PROD_HOST", "slave.db.yallasvc.net")

    return {
        "host": host,
        "port": int(os.environ.get("PG_PROD_PORT", "5432")),
        "user": os.environ.get("PG_PROD_USER", ""),
        "password": os.environ.get("PG_PROD_PASSWORD", ""),
        "dbname": db,
        "sslmode": "require",
    }


MAX_RESULT_CHARS = 50_000
STATEMENT_TIMEOUT_MS = 30_000
ALLOWED_PATH_ROOT = "/zeroclaw-data/workspaces/"
ALLOWED_PATH_SEGMENT = "/workspace/state/sql/"

_RETRYABLE_PG_PATTERNS = (
    "timeout expired",
    "could not connect",
    "connection refused",
    "server closed the connection",
    "server closed the connection unexpectedly",
    "the database system is starting up",
    "no route to host",
    "network is unreachable",
)
RETRY_MAX_ATTEMPTS = 3
RETRY_BACKOFFS_SECS = (0.5, 1.5)
_NON_RETRYABLE_SQLSTATES = frozenset({"57014"})
_DML_RE = re.compile(
    r"\b(INSERT|UPDATE|DELETE|MERGE|TRUNCATE)\b",
    re.IGNORECASE,
)

PG_HOST_ROUTE_TARGETS = {
    "slave.db.yallasvc.net": "46.4.211.98",
    "slave.catalog.db.yallasvc.net": "195.201.82.13",
}
VPN_ROUTE_WAIT_SECS = int(os.environ.get("MCP_DB_VPN_ROUTE_WAIT_SECS", "25"))
VPN_ROUTE_POLL_SECS = float(os.environ.get("MCP_DB_VPN_ROUTE_POLL_SECS", "1"))


def _is_transient_pg_error(exc: BaseException) -> bool:
    """Return True for PG failures that are safe to retry for read-only SQL."""
    sqlstate = getattr(exc, "pgcode", None)
    if sqlstate in _NON_RETRYABLE_SQLSTATES:
        return False
    if isinstance(exc, pg_errors.QueryCanceled):
        return False
    msg = str(exc).lower()
    return any(pattern in msg for pattern in _RETRYABLE_PG_PATTERNS)


def _sql_is_retry_safe(sql: str) -> bool:
    """True iff SQL is a pure SELECT/WITH read query without nested DML."""
    cleaned = _strip_sql_noise(sql).strip().lstrip("(").lstrip()
    first_word = cleaned.split()[0].upper() if cleaned else ""
    if first_word not in ("SELECT", "WITH"):
        return False
    return _DML_RE.search(cleaned) is None


def _route_uses_tun0(ip: str) -> bool:
    """Check whether Linux routes an IP through tun0; skip where unavailable."""
    if sys.platform == "darwin" or shutil.which("ip") is None:
        return True
    try:
        proc = subprocess.run(
            ["ip", "route", "get", ip],
            capture_output=True,
            text=True,
            timeout=2,
            check=False,
        )
    except FileNotFoundError:
        return True
    except subprocess.SubprocessError:
        return False
    except Exception:
        return False
    return proc.returncode == 0 and " dev tun0 " in f" {proc.stdout} "


def _wait_for_vpn_route(host: str) -> None:
    """Wait until a known prod host route goes through tun0."""
    ip = PG_HOST_ROUTE_TARGETS.get(host)
    if not ip:
        return
    deadline = time.monotonic() + VPN_ROUTE_WAIT_SECS
    while time.monotonic() < deadline:
        if _route_uses_tun0(ip):
            return
        time.sleep(VPN_ROUTE_POLL_SECS)
    raise RuntimeError(f"vpn_unavailable: route to {host} ({ip}) is not via tun0")


def _execute_pg_with_retry(
    connect_kwargs: dict,
    sql: str,
    statement_timeout_ms: int,
) -> tuple[str, object, object]:
    """Run PG SQL, retrying transient connect/execute failures for reads only."""
    retry_safe = _sql_is_retry_safe(sql)
    last_exc: BaseException | None = None
    host = connect_kwargs.get("host")
    for attempt in range(RETRY_MAX_ATTEMPTS):
        conn = None
        is_dml = False
        try:
            if host:
                _wait_for_vpn_route(host)
            conn = psycopg2.connect(**connect_kwargs, connect_timeout=10)
            with conn.cursor() as cur:
                cur.execute(f"SET statement_timeout = {statement_timeout_ms}")
                cur.execute(sql)
                if cur.description is None:
                    is_dml = True
                    return ("dml", cur.rowcount, conn)
                columns = [desc[0] for desc in cur.description]
                rows = cur.fetchall()
                return ("rows", columns, rows)
        except psycopg2.Error as exc:
            last_exc = exc
            if (
                not retry_safe
                or not _is_transient_pg_error(exc)
                or attempt == RETRY_MAX_ATTEMPTS - 1
            ):
                raise
            print(
                f"[retry] pg execute attempt {attempt + 1} failed: {exc}",
                file=sys.stderr,
                flush=True,
            )
            time.sleep(RETRY_BACKOFFS_SECS[attempt])
        finally:
            if conn is not None and not is_dml:
                try:
                    conn.close()
                except Exception:
                    pass
    raise last_exc or RuntimeError("unreachable pg retry state")


_RETRYABLE_CH_EXC = (
    urllib.error.URLError,
    socket.timeout,
    http.client.RemoteDisconnected,
)


def _ch_post_with_retry(req: urllib.request.Request, timeout: int) -> bytes:
    """POST ClickHouse query with retry for full-buffer idempotent reads."""
    last_exc: BaseException | None = None
    for attempt in range(RETRY_MAX_ATTEMPTS):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except urllib.error.HTTPError:
            raise
        except _RETRYABLE_CH_EXC as exc:
            last_exc = exc
            if attempt == RETRY_MAX_ATTEMPTS - 1:
                raise
            print(
                f"[retry] ch attempt {attempt + 1} failed: {exc}",
                file=sys.stderr,
                flush=True,
            )
            time.sleep(RETRY_BACKOFFS_SECS[attempt])
    raise last_exc or RuntimeError("unreachable ch retry state")


def _ch_post_with_retry_initial(req: urllib.request.Request, timeout: int):
    """Open a ClickHouse stream with retry before the first byte only."""
    last_exc: BaseException | None = None
    for attempt in range(RETRY_MAX_ATTEMPTS):
        try:
            return urllib.request.urlopen(req, timeout=timeout)
        except urllib.error.HTTPError:
            raise
        except _RETRYABLE_CH_EXC as exc:
            last_exc = exc
            if attempt == RETRY_MAX_ATTEMPTS - 1:
                raise
            print(
                f"[retry] ch stream initial attempt {attempt + 1} failed: {exc}",
                file=sys.stderr,
                flush=True,
            )
            time.sleep(RETRY_BACKOFFS_SECS[attempt])
    raise last_exc or RuntimeError("unreachable ch stream retry state")


def _count_openvpn_resets(log_path: str, window_secs: int = 300) -> int:
    """Count recent-looking OpenVPN soft reconnects from the tail of its log."""
    try:
        with open(log_path, "rb") as f:
            f.seek(0, os.SEEK_END)
            size = f.tell()
            f.seek(max(0, size - 256_000))
            tail = f.read().decode("utf-8", errors="replace")
    except OSError:
        return 0
    # OpenVPN logs are not reliably machine-parseable across configs. Keep this
    # cheap and deterministic; smoke checks divide by their observed window.
    _ = window_secs
    return tail.count("SIGUSR1[soft,connection-reset]")


def validate_file_path(path: str) -> str:
    """Validate that a file path is absolute and under a per-workspace SQL directory."""
    if not os.path.isabs(path):
        raise ValueError(f"Path must be absolute, got: '{path}'")
    real = os.path.realpath(path)
    if not real.startswith(ALLOWED_PATH_ROOT):
        raise ValueError(
            f"Path '{path}' resolves to '{real}' which is outside "
            f"the allowed root '{ALLOWED_PATH_ROOT}'"
        )
    if ALLOWED_PATH_SEGMENT not in real or not real.endswith(".sql"):
        raise ValueError(
            "SQL file path must live under '<workspace>/state/sql/' and end with '.sql'"
        )
    return real


def truncate_result(data: str, max_chars: int = MAX_RESULT_CHARS) -> tuple[str, bool]:
    """Truncate result string if it exceeds max_chars. Returns (data, was_truncated)."""
    if len(data) <= max_chars:
        return data, False
    warning = f"\n\n[WARNING: Result truncated from {len(data)} to {max_chars} characters]"
    return data[:max_chars] + warning, True


def execute_query(env: str, db: str, sql: str) -> dict:
    """Execute SQL query and return structured result."""
    check_sql_safety(sql)
    sql = apply_auto_limit(sql)
    params = resolve_connection_params(env, db)

    kind, payload_a, payload_b = _execute_pg_with_retry(
        params, sql, STATEMENT_TIMEOUT_MS
    )
    if kind == "dml":
        row_count = payload_a
        conn = payload_b
        try:
            conn.commit()
        finally:
            conn.close()
        return {
            "columns": [],
            "rows": [],
            "row_count": row_count,
            "message": f"{row_count} row(s) affected",
        }

    columns = payload_a
    rows = payload_b
    safe_rows = []
    for row in rows:
        safe_rows.append([
            str(v) if v is not None and not isinstance(v, (int, float, bool, str)) else v
            for v in row
        ])
    return {
        "columns": columns,
        "rows": safe_rows,
        "row_count": len(safe_rows),
    }


# ---------------------------------------------------------------------------
# ClickHouse
# ---------------------------------------------------------------------------

CH_DML_BLOCKLIST_RE = re.compile(
    r"\b(INSERT|UPDATE|DELETE|DROP|TRUNCATE|ALTER|CREATE|GRANT|REVOKE|"
    r"OPTIMIZE|SYSTEM|ATTACH|DETACH)\b",
    re.IGNORECASE,
)

CH_FORMAT_RE = re.compile(r"\bFORMAT\s+\w+", re.IGNORECASE)
CH_IDENTIFIER_RE = re.compile(r"^[a-zA-Z_][a-zA-Z0-9_]*$")


def check_ch_sql_safety(sql: str) -> None:
    """Raise ValueError if ClickHouse SQL contains blocked statements."""
    cleaned = _strip_sql_noise(sql)
    first_word = cleaned.strip().split()[0].upper() if cleaned.strip() else ""
    if first_word in ("SHOW", "DESCRIBE", "DESC", "EXPLAIN"):
        return
    match = CH_DML_BLOCKLIST_RE.search(cleaned)
    if match:
        raise ValueError(
            f"Statement blocked for ClickHouse: '{match.group()}'. "
            f"Only SELECT queries are allowed (read-only prod access)."
        )


def _ch_connection_params() -> dict[str, str]:
    return {
        "host": os.environ.get("CH_HOST", "ch-prod.yallasvc.net"),
        "port": os.environ.get("CH_PORT", "8123"),
        "user": os.environ.get("CH_USER", "app_data_uploader_20250626"),
        "password": os.environ.get(
            "CH_PASSWORD", "XTGBs7JjcCT5XiT3L8d8HYc7KQdpAk0t"
        ),
    }


def _validate_ch_identifier(name: str) -> str:
    """Validate ClickHouse identifier (database/table name)."""
    if not CH_IDENTIFIER_RE.match(name):
        raise ValueError(
            f"Invalid identifier: '{name}'. "
            f"Only letters, digits, and underscores are allowed."
        )
    return name


def execute_ch_query(sql: str, db: str | None = None) -> str:
    """Execute ClickHouse query via HTTP API. Returns raw text response."""
    check_ch_sql_safety(sql)

    stripped = sql.strip().rstrip(";")

    # Auto-limit SELECT queries
    upper = stripped.upper()
    if (upper.startswith("SELECT") or upper.startswith("WITH")) and not LIMIT_RE.search(
        stripped
    ):
        stripped += " LIMIT 1000"

    # Auto-add FORMAT TSVWithNames
    if not CH_FORMAT_RE.search(stripped):
        stripped += "\nFORMAT TSVWithNames"

    params = _ch_connection_params()
    qs_params: dict[str, str] = {
        "user": params["user"],
        "password": params["password"],
    }
    if db:
        qs_params["database"] = db
    url = f"http://{params['host']}:{params['port']}/?{urllib.parse.urlencode(qs_params)}"

    req = urllib.request.Request(url, data=stripped.encode("utf-8"), method="POST")
    try:
        return _ch_post_with_retry(req, timeout=30).decode("utf-8")
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise ValueError(
            f"ClickHouse error (HTTP {e.code}): {body}"
        ) from None


# ---------------------------------------------------------------------------
# Export validators
# ---------------------------------------------------------------------------

EXPORT_FORMATS = frozenset({"csv", "json", "xlsx"})
OUT_NAME_RE = re.compile(r"^[A-Za-z0-9_\-]{1,64}$")

MAX_EXPORT_ROWS = 500_000
MAX_EXPORT_TIMEOUT_SECS = 300


def validate_out_name(name: str) -> str:
    """Validate human-readable base name for exported file."""
    if not isinstance(name, str) or not OUT_NAME_RE.match(name):
        raise ValueError(
            f"Invalid out_name: must match [A-Za-z0-9_-]{{1,64}}, got {name!r}"
        )
    return name


def validate_export_format(fmt: str) -> str:
    """Validate export format and normalize to lowercase."""
    if not isinstance(fmt, str):
        raise ValueError(f"Invalid format: must be str, got {type(fmt).__name__}")
    normalized = fmt.strip().lower()
    if normalized not in EXPORT_FORMATS:
        raise ValueError(
            f"Invalid format {fmt!r}: must be one of {sorted(EXPORT_FORMATS)}"
        )
    return normalized


def validate_export_bounds(max_rows: int, timeout_secs: int) -> tuple[int, int]:
    """Validate row/timeout bounds for export."""
    if not isinstance(max_rows, int) or max_rows <= 0 or max_rows > MAX_EXPORT_ROWS:
        raise ValueError(
            f"Invalid max_rows={max_rows}: must be 1..{MAX_EXPORT_ROWS}"
        )
    if (
        not isinstance(timeout_secs, int)
        or timeout_secs <= 0
        or timeout_secs > MAX_EXPORT_TIMEOUT_SECS
    ):
        raise ValueError(
            f"Invalid timeout_secs={timeout_secs}: must be 1..{MAX_EXPORT_TIMEOUT_SECS}"
        )
    return max_rows, timeout_secs


def resolve_upload_path(user_workspace: str, out_name: str, fmt: str) -> str:
    """Build absolute path under <user_workspace>/uploads/ with uuid prefix.

    Validates that the result stays under ALLOWED_PATH_ROOT (per-user workspace).
    Returns the absolute path.
    """
    import uuid as _uuid

    real_workspace = os.path.realpath(user_workspace)
    if not real_workspace.startswith(ALLOWED_PATH_ROOT):
        raise ValueError(
            f"Workspace path {user_workspace!r} resolves to {real_workspace!r} "
            f"outside allowed root {ALLOWED_PATH_ROOT!r}"
        )
    uploads_dir = os.path.join(real_workspace, "uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    name = validate_out_name(out_name)
    extension = validate_export_format(fmt)
    short = _uuid.uuid4().hex[:12]
    return os.path.join(uploads_dir, f"{short}-{name}.{extension}")


# ---------------------------------------------------------------------------
# Format writers
# ---------------------------------------------------------------------------

import csv as _csv  # noqa: E402
from typing import Iterator  # noqa: E402

CSV_PREVIEW_VALUE_MAX = 200


def _coerce_csv_value(v: object) -> str:
    if v is None:
        return ""
    return str(v)


def write_csv_stream(
    out_path: str,
    columns: list[str],
    row_source: Iterator[tuple],
    max_rows: int,
) -> tuple[int, bool]:
    """Stream rows to a CSV file. Returns (rows_written, truncated)."""
    written = 0
    truncated = False
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        writer = _csv.writer(fh, quoting=_csv.QUOTE_MINIMAL)
        writer.writerow(columns)
        for row in row_source:
            if written >= max_rows:
                truncated = True
                break
            writer.writerow([_coerce_csv_value(v) for v in row])
            written += 1
    return written, truncated


def read_csv_preview(path: str, limit: int = 5) -> list[dict]:
    """Read first `limit` data rows from CSV as list of dicts."""
    with open(path, newline="", encoding="utf-8") as fh:
        reader = _csv.DictReader(fh)
        out: list[dict] = []
        for row in reader:
            if len(out) >= limit:
                break
            trimmed = {
                k: (v[:CSV_PREVIEW_VALUE_MAX] if isinstance(v, str) else v)
                for k, v in row.items()
            }
            out.append(trimmed)
        return out


# ---------------------------------------------------------------------------
# execute_query_to_file (PostgreSQL)
# ---------------------------------------------------------------------------

SELECT_ONLY_RE = re.compile(r"^\s*(SELECT|WITH)\b", re.IGNORECASE)


def _check_select_only(sql: str) -> None:
    cleaned = _strip_sql_noise(sql).strip()
    if not SELECT_ONLY_RE.match(cleaned):
        raise ValueError(
            "query_to_file accepts only SELECT/WITH queries. "
            "For DML use query_file."
        )


def _open_pg_connection(params: dict, timeout_secs: int):
    """Open a PG connection with retry on initial transient connect failures."""
    _ = timeout_secs
    host = params.get("host")
    last_exc: BaseException | None = None
    for attempt in range(RETRY_MAX_ATTEMPTS):
        try:
            if host:
                _wait_for_vpn_route(host)
            return psycopg2.connect(
                host=params["host"],
                port=params["port"],
                user=params["user"],
                password=params["password"],
                dbname=params["dbname"],
                sslmode=params["sslmode"],
                connect_timeout=10,
            )
        except psycopg2.OperationalError as exc:
            last_exc = exc
            if not _is_transient_pg_error(exc) or attempt == RETRY_MAX_ATTEMPTS - 1:
                raise
            print(
                f"[retry] pg connect attempt {attempt + 1} failed: {exc}",
                file=sys.stderr,
                flush=True,
            )
            time.sleep(RETRY_BACKOFFS_SECS[attempt])
    raise last_exc or RuntimeError("unreachable pg connect retry state")


def execute_query_to_file(
    env: str,
    db: str,
    sql: str,
    user_workspace: str,
    out_name: str,
    format: str,
    max_rows: int,
    timeout_secs: int,
) -> dict:
    """Stream SQL results to a file under <user_workspace>/uploads/.

    Returns a metadata dict: path, absolute_path, format, row_count, column_names,
    size_bytes, duration_ms, preview (first 5 rows), truncated, preview_capped_at.
    """
    import time as _time
    fmt = validate_export_format(format)
    max_rows, timeout_secs = validate_export_bounds(max_rows, timeout_secs)
    check_sql_safety(sql)
    _check_select_only(sql)
    if fmt == "xlsx":
        _ensure_openpyxl()
    abs_path = resolve_upload_path(user_workspace, out_name, fmt)
    params = resolve_connection_params(env, db)

    start = _time.monotonic()
    conn = _open_pg_connection(params, timeout_secs)
    columns: list[str] = []
    rows_written = 0
    truncated = False
    try:
        # statement_timeout must be applied on a plain cursor BEFORE opening a
        # server-side named cursor — psycopg2 cannot run SET inside a named cursor.
        with conn.cursor() as setup_cur:
            setup_cur.execute(f"SET statement_timeout = {timeout_secs * 1000}")
        with conn.cursor(name=f"export_{os.path.basename(abs_path)}") as cur:
            cur.itersize = 5000
            cur.execute(sql)
            # Named (server-side) cursors only populate cur.description after
            # the first FETCH. Peek one row to force binding, then prepend it
            # to the iterator.
            first_row = cur.fetchone()
            columns = [d[0] for d in cur.description] if cur.description else []

            def _source():
                if first_row is not None:
                    yield first_row
                yield from cur

            if fmt == "csv":
                rows_written, truncated = write_csv_stream(
                    abs_path, columns, _source(), max_rows
                )
            elif fmt == "json":
                rows_written, truncated = write_json_stream(
                    abs_path, columns, _source(), max_rows
                )
            else:  # xlsx
                rows_written, truncated = write_xlsx(
                    abs_path, columns, _source(), max_rows
                )
    finally:
        conn.close()

    duration_ms = int((_time.monotonic() - start) * 1000)
    size_bytes = os.path.getsize(abs_path) if os.path.isfile(abs_path) else 0
    preview = _read_preview(abs_path, fmt, limit=5)

    rel_path = os.path.relpath(abs_path, start=os.path.realpath(user_workspace))
    return {
        "path": rel_path,
        "absolute_path": abs_path,
        "format": fmt,
        "row_count": rows_written,
        "column_names": columns,
        "size_bytes": size_bytes,
        "duration_ms": duration_ms,
        "preview": preview,
        "preview_capped_at": len(preview),
        "truncated": truncated,
    }


def _read_preview(path: str, fmt: str, limit: int) -> list[dict]:
    if fmt == "csv":
        return read_csv_preview(path, limit=limit)
    if fmt == "json":
        return read_json_preview(path, limit=limit)
    return read_xlsx_preview(path, limit=limit)


def _ensure_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise ValueError(
            "xlsx format requires 'openpyxl' package which is not installed"
        ) from e


def write_json_stream(
    out_path: str,
    columns: list[str],
    row_source: Iterator[tuple],
    max_rows: int,
) -> tuple[int, bool]:
    """Stream rows as NDJSON (one JSON object per line)."""
    written = 0
    truncated = False
    with open(out_path, "w", encoding="utf-8") as fh:
        for row in row_source:
            if written >= max_rows:
                truncated = True
                break
            obj = {col: _coerce_json_value(v) for col, v in zip(columns, row)}
            fh.write(json.dumps(obj, ensure_ascii=False, default=str) + "\n")
            written += 1
    return written, truncated


def _coerce_json_value(v: object):
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


def read_json_preview(path: str, limit: int = 5) -> list[dict]:
    out: list[dict] = []
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            if len(out) >= limit:
                break
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return out


XLSX_HARD_CAP = 100_000


def write_xlsx(
    out_path: str,
    columns: list[str],
    row_source: Iterator[tuple],
    max_rows: int,
) -> tuple[int, bool]:
    """Write rows to XLSX. Enforces min(max_rows, XLSX_HARD_CAP) due to Excel limits."""
    import openpyxl
    from openpyxl.utils.exceptions import IllegalCharacterError

    effective_cap = min(max_rows, XLSX_HARD_CAP)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="export")
    ws.append(columns)
    written = 0
    truncated = False
    for row in row_source:
        if written >= effective_cap:
            truncated = True
            break
        try:
            ws.append([_coerce_xlsx_value(v) for v in row])
        except IllegalCharacterError:
            ws.append([_strip_xlsx_illegal(_coerce_xlsx_value(v)) for v in row])
        written += 1
    wb.save(out_path)
    # If we hit the effective cap but source has more rows, we may have consumed
    # one extra row from the iterator; that's OK — truncated=True is accurate.
    return written, truncated


_XLSX_ILLEGAL_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def _coerce_xlsx_value(v: object):
    if v is None or isinstance(v, (int, float, bool, str)):
        return v
    return str(v)


def _strip_xlsx_illegal(v: object):
    if isinstance(v, str):
        return _XLSX_ILLEGAL_RE.sub("", v)
    return v


def read_xlsx_preview(path: str, limit: int = 5) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    out: list[dict] = []
    for row in rows_iter:
        if len(out) >= limit:
            break
        out.append({h: v for h, v in zip(header, row)})
    wb.close()
    return out


# ---------------------------------------------------------------------------
# execute_ch_query_to_file (ClickHouse)
# ---------------------------------------------------------------------------


def execute_ch_query_to_file(
    sql: str,
    user_workspace: str,
    out_name: str,
    format: str = "csv",
    db: str = "analytics",
    max_rows: int = 100_000,
    timeout_secs: int = 120,
) -> dict:
    """Stream ClickHouse SELECT results to a file under <user_workspace>/uploads/."""
    import time as _time
    fmt = validate_export_format(format)
    max_rows, timeout_secs = validate_export_bounds(max_rows, timeout_secs)
    check_ch_sql_safety(sql)
    _check_select_only(sql)
    if fmt == "xlsx":
        _ensure_openpyxl()
    abs_path = resolve_upload_path(user_workspace, out_name, fmt)

    # ClickHouse native format = CSVWithNames (always); xlsx path pipes
    # CSVWithNames through pandas.
    ch_format = "CSVWithNames" if fmt in ("csv", "xlsx") else "JSONEachRow"
    stripped = sql.strip().rstrip(";")
    stripped = CH_FORMAT_RE.sub("", stripped).strip()
    stripped += (
        f"\nSETTINGS max_result_rows = {max_rows},"
        f" result_overflow_mode = 'break'"
        f"\nFORMAT {ch_format}"
    )

    params = _ch_connection_params()
    qs: dict[str, str] = {"user": params["user"], "password": params["password"]}
    if db:
        qs["database"] = db
    url = f"http://{params['host']}:{params['port']}/?{urllib.parse.urlencode(qs)}"
    req = urllib.request.Request(
        url, data=stripped.encode("utf-8"), method="POST"
    )

    start = _time.monotonic()
    tmp_path = abs_path if fmt != "xlsx" else abs_path + ".tmp.csv"
    try:
        resp = _ch_post_with_retry_initial(req, timeout=timeout_secs)
        try:
            with open(tmp_path, "wb") as fh:
                while True:
                    chunk = resp.read(65536)
                    if not chunk:
                        break
                    fh.write(chunk)
        finally:
            close = getattr(resp, "close", None)
            if close is not None:
                close()
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        raise ValueError(f"ClickHouse HTTP {e.code}: {body}") from None
    except (OSError, socket.error):
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise

    # Parse CSV/JSON header + row_count; convert to xlsx if needed.
    if fmt == "csv":
        columns, row_count, truncated = _summarize_csv_file(tmp_path, max_rows)
    elif fmt == "json":
        columns, row_count, truncated = _summarize_json_file(tmp_path, max_rows)
    else:  # xlsx
        columns, row_count, truncated = _convert_csv_to_xlsx(tmp_path, abs_path, max_rows)
        os.unlink(tmp_path)

    duration_ms = int((_time.monotonic() - start) * 1000)
    size_bytes = os.path.getsize(abs_path)
    preview = _read_preview(abs_path, fmt, limit=5)
    rel_path = os.path.relpath(abs_path, start=os.path.realpath(user_workspace))
    return {
        "path": rel_path,
        "absolute_path": abs_path,
        "format": fmt,
        "row_count": row_count,
        "column_names": columns,
        "size_bytes": size_bytes,
        "duration_ms": duration_ms,
        "preview": preview,
        "preview_capped_at": len(preview),
        "truncated": truncated,
    }


def _summarize_csv_file(path: str, max_rows: int) -> tuple[list[str], int, bool]:
    with open(path, newline="", encoding="utf-8") as fh:
        reader = _csv.reader(fh)
        try:
            columns = next(reader)
        except StopIteration:
            return [], 0, False
        count = sum(1 for _ in reader)
    truncated = count >= max_rows
    return columns, count, truncated


def _summarize_json_file(path: str, max_rows: int) -> tuple[list[str], int, bool]:
    columns: list[str] = []
    count = 0
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            if not columns:
                try:
                    columns = list(json.loads(line).keys())
                except json.JSONDecodeError:
                    pass
            count += 1
    return columns, count, count >= max_rows


def _convert_csv_to_xlsx(
    csv_path: str, xlsx_path: str, max_rows: int
) -> tuple[list[str], int, bool]:
    """Convert a CSVWithNames file to xlsx, enforcing XLSX_HARD_CAP."""
    import openpyxl
    effective_cap = min(max_rows, XLSX_HARD_CAP)
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title="export")
    columns: list[str] = []
    written = 0
    truncated = False
    with open(csv_path, newline="", encoding="utf-8") as fh:
        reader = _csv.reader(fh)
        try:
            columns = next(reader)
            ws.append(columns)
        except StopIteration:
            wb.save(xlsx_path)
            return [], 0, False
        for row in reader:
            if written >= effective_cap:
                truncated = True
                break
            ws.append(row)
            written += 1
    wb.save(xlsx_path)
    return columns, written, truncated


# ---------------------------------------------------------------------------
# MCP Server
# ---------------------------------------------------------------------------

from mcp.server.fastmcp import FastMCP  # noqa: E402

mcp = FastMCP(
    "lalafo-db",
    stateless_http=True,
    json_response=True,
    streamable_http_path="/",
)


@mcp.tool()
def query(env: str, db: str, sql: str) -> str:
    """Execute a SQL query against a Lalafo PostgreSQL database.

    Args:
        env: Environment — "stage" (full access) or "prod" (read-only replica).
        db: Database name, e.g. "catalog_microservice", "user_microservice".
        sql: SQL query string. DDL is blocked. SELECT auto-limited to 1000 rows.
    """
    try:
        result = execute_query(env, db, sql)
        output = json.dumps(result, ensure_ascii=False, default=str)
        output, _ = truncate_result(output)
        return output
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@mcp.tool()
def query_file(env: str, db: str, sql_file: str) -> str:
    """Execute SQL from a file against a Lalafo PostgreSQL database.

    Args:
        env: Environment — "stage" or "prod".
        db: Database name.
        sql_file: Absolute path to SQL file under <workspace>/state/sql/*.sql.
    """
    try:
        real_path = validate_file_path(sql_file)
        if not os.path.isfile(real_path):
            return json.dumps({"error": f"File not found: {sql_file}"})
        with open(real_path) as f:
            sql = f.read()
        if not sql.strip():
            return json.dumps({"error": "SQL file is empty"})
        result = execute_query(env, db, sql)
        output = json.dumps(result, ensure_ascii=False, default=str)
        output, _ = truncate_result(output)
        return output
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@mcp.tool()
def databases(env: str) -> str:
    """List available PostgreSQL databases and their descriptions.

    Args:
        env: Environment — "stage" or "prod".
    """
    if env not in ("stage", "prod"):
        return json.dumps({"error": f"Invalid environment: '{env}'. Must be 'stage' or 'prod'."})
    db_list = [
        {"name": name, "description": info["description"]}
        for name, info in sorted(DATABASES.items())
    ]
    return json.dumps(db_list, ensure_ascii=False)


@mcp.tool()
def ch_query(sql: str, db: str = "analytics") -> str:
    """Execute a read-only SQL query against Lalafo ClickHouse (prod).

    ClickHouse contains analytics tables and pg_* replicas of all PostgreSQL
    microservices. Cross-service JOINs are possible using full db.table names
    (e.g. pg_catalog_microservice.ad JOIN pg_user_microservice.user).

    Args:
        sql: SQL query. Only SELECT allowed. Auto-limited to 1000 rows.
             FORMAT TSVWithNames added automatically.
        db: ClickHouse database, default "analytics". For cross-db JOINs
            use "default" and qualify tables with full db.table names.
    """
    try:
        result = execute_ch_query(sql, db)
        result, _ = truncate_result(result)
        return result
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@mcp.tool()
def ch_databases() -> str:
    """List available ClickHouse databases (analytics, pg_* replicas, etc.)."""
    try:
        result = execute_ch_query("SHOW DATABASES")
        return result
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@mcp.tool()
def ch_tables(db: str) -> str:
    """List tables in a ClickHouse database.

    Args:
        db: Database name, e.g. "analytics", "pg_catalog_microservice".
    """
    try:
        safe_db = _validate_ch_identifier(db)
        result = execute_ch_query(f"SHOW TABLES FROM {safe_db}")
        return result
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@mcp.tool()
def query_to_file(
    env: str,
    db: str,
    sql: str,
    workspace: str,
    out_name: str,
    format: str = "csv",
    max_rows: int = 100_000,
    timeout_secs: int = 120,
) -> str:
    """Execute SELECT and stream results to a file in workspace/uploads/.

    Returns metadata (path, row_count, column_names, preview of first 5 rows).
    The full result is in the file — NEVER loaded into agent context. Use this
    tool when the user asks to export/download data or the result has more than
    ~100 rows.

    Args:
        env: "stage" | "prod".
        db: Database name (e.g. "micromarket_microservice").
        sql: SELECT/WITH query. DDL/DML blocked.
        workspace: Absolute path of the caller's workspace directory
            (the one that contains uploads/, state/, memory/). E.g.
            /zeroclaw-data/workspaces/tg_<uid>/workspace
        out_name: Base name for output file, [A-Za-z0-9_-]{1,64}.
        format: "csv" (default) | "json" | "xlsx".
        max_rows: Row limit, default 100_000, hard cap 500_000.
        timeout_secs: Query timeout, default 120, hard cap 300.
    """
    try:
        result = execute_query_to_file(
            env=env, db=db, sql=sql,
            user_workspace=workspace,
            out_name=out_name, format=format,
            max_rows=max_rows, timeout_secs=timeout_secs,
        )
        return json.dumps(result, ensure_ascii=False, default=str)
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@mcp.tool()
def ch_query_to_file(
    sql: str,
    workspace: str,
    out_name: str,
    format: str = "csv",
    db: str = "analytics",
    max_rows: int = 100_000,
    timeout_secs: int = 120,
) -> str:
    """Execute ClickHouse SELECT and stream results to a file in workspace/uploads/.

    See `query_to_file` for usage. ClickHouse supports cross-service JOINs via
    pg_* replicas (e.g. pg_catalog_microservice.ad JOIN pg_user_microservice.user).

    Args:
        sql: SELECT/WITH query. DDL/DML blocked.
        workspace: Absolute path of the caller's workspace directory
            (the one that contains uploads/, state/, memory/). E.g.
            /zeroclaw-data/workspaces/tg_<uid>/workspace
        out_name: Base name for output file, [A-Za-z0-9_-]{1,64}.
        format: "csv" (default) | "json" | "xlsx".
        db: ClickHouse database, default "analytics".
        max_rows: Row limit, default 100_000, hard cap 500_000.
        timeout_secs: Query timeout, default 120, hard cap 300.
    """
    try:
        result = execute_ch_query_to_file(
            sql=sql, user_workspace=workspace,
            out_name=out_name, format=format, db=db,
            max_rows=max_rows, timeout_secs=timeout_secs,
        )
        return json.dumps(result, ensure_ascii=False, default=str)
    except Exception as e:
        return json.dumps({"error": str(e)}, ensure_ascii=False)


@mcp.tool()
def health() -> str:
    """Check VPN connectivity, PostgreSQL and ClickHouse reachability."""
    result: dict = {
        "status": "ok",
        "vpn": "unknown",
        "stage": "unknown",
        "prod": "unknown",
        "prod_catalog": "unknown",
        "clickhouse": "unknown",
    }

    # Check tun0
    try:
        output = subprocess.run(
            ["ip", "link", "show", "tun0"],
            capture_output=True, text=True, timeout=5,
        )
        result["vpn"] = "tun0 up" if output.returncode == 0 else "tun0 down"
    except Exception:
        result["vpn"] = "tun0 check failed"

    # Check stage host TCP
    stage_host = os.environ.get("PG_STAGE_HOST", "stage.postgres.kube-two.yallasvc.net")
    stage_port = int(os.environ.get("PG_STAGE_PORT", "6432"))
    try:
        sock = socket.create_connection((stage_host, stage_port), timeout=2)
        sock.close()
        result["stage"] = "reachable"
    except Exception:
        result["stage"] = "unreachable"

    # Check prod host TCP
    prod_host = os.environ.get("PG_PROD_HOST", "slave.db.yallasvc.net")
    prod_port = int(os.environ.get("PG_PROD_PORT", "5432"))
    try:
        sock = socket.create_connection((prod_host, prod_port), timeout=2)
        sock.close()
        result["prod"] = "reachable"
    except Exception:
        result["prod"] = "unreachable"

    # Check prod catalog host TCP
    prod_catalog_host = os.environ.get("PG_PROD_CATALOG_HOST", "slave.catalog.db.yallasvc.net")
    try:
        sock = socket.create_connection((prod_catalog_host, prod_port), timeout=2)
        sock.close()
        result["prod_catalog"] = "reachable"
    except Exception:
        result["prod_catalog"] = "unreachable"

    # Check ClickHouse HTTP
    ch_params = _ch_connection_params()
    try:
        ch_url = f"http://{ch_params['host']}:{ch_params['port']}/ping"
        ch_req = urllib.request.Request(ch_url, method="GET")
        with urllib.request.urlopen(ch_req, timeout=5):
            result["clickhouse"] = "reachable"
    except Exception:
        result["clickhouse"] = "unreachable"

    result["routes"] = {
        "prod": "tun0" if _route_uses_tun0(
            PG_HOST_ROUTE_TARGETS["slave.db.yallasvc.net"]
        ) else "off-vpn",
        "prod_catalog": "tun0" if _route_uses_tun0(
            PG_HOST_ROUTE_TARGETS["slave.catalog.db.yallasvc.net"]
        ) else "off-vpn",
    }
    result["openvpn_recent_resets"] = _count_openvpn_resets(
        os.environ.get("OPENVPN_LOG_PATH", "/tmp/openvpn.log"),
        window_secs=300,
    )

    if (
        result["vpn"] != "tun0 up"
        or result["stage"] != "reachable"
        or result["prod"] != "reachable"
        or result["prod_catalog"] != "reachable"
        or result["clickhouse"] != "reachable"
    ):
        result["status"] = "degraded"
    return json.dumps(result)


# ---------------------------------------------------------------------------
# HTTP app: /mcp (MCP protocol) + /health (readiness probe)
# ---------------------------------------------------------------------------

def create_app():
    """Create Starlette app with both /mcp (MCP protocol) and /health (HTTP probe)."""
    import contextlib
    from collections.abc import AsyncIterator

    from starlette.applications import Starlette
    from starlette.responses import JSONResponse
    from starlette.routing import Mount, Route

    async def health_handler(request):
        result = json.loads(health())
        status = 200 if result.get("status") == "ok" else 503
        return JSONResponse(result, status_code=status)

    @contextlib.asynccontextmanager
    async def lifespan(app: Starlette) -> AsyncIterator[None]:
        async with mcp.session_manager.run():
            yield

    return Starlette(
        routes=[
            Route("/health", health_handler, methods=["GET"]),
            Mount("/mcp", app=mcp.streamable_http_app()),
        ],
        lifespan=lifespan,
    )


if __name__ == "__main__":
    import uvicorn

    port = int(os.environ.get("MCP_DB_PORT", "4000"))
    print(f"[mcp-db-server] Starting on port {port}")
    uvicorn.run(create_app(), host="0.0.0.0", port=port, log_level="warning")
